#!/usr/bin/env python3
"""
Converte Skiathos.xlsx (foglio "Giornaliero") in webapp/data/itinerary.json.

Rilancialo ogni volta che modifichi l'Excel:
    python3 scripts/xlsx_to_json.py

Logica di lettura (il foglio non ha una griglia rigida: ogni giorno è una
colonna B..M). Per ogni colonna/giorno raccogliamo tutti i blocchi di testo
non vuoti dalle righe 4-11 (attività di giorno), in ordine, e li
classifichiamo così:
  - "heading": testo corto e in grassetto/font grande -> titolo/luogo
  - "paragraph": testo lungo (>=100 caratteri) -> descrizione della giornata
  - "note": tutto il resto -> chip con icona (voli, bagagli, spesa...)

La riga 12 (cena) e la riga 13 (nota/attività serale) sono sempre nella
stessa posizione per ogni giorno, quindi le trattiamo separatamente come
sezione "Sera" (vedi classify_dinner).
"""
import hashlib
import html
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path
from urllib.parse import quote

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
XLSX_PATH = ROOT / "Skiathos.xlsx"
WEBAPP_DIR = Path(__file__).resolve().parent.parent
OUT_PATH = WEBAPP_DIR / "data" / "itinerary.json"
SW_PATH = WEBAPP_DIR / "sw.js"

SHEET_NAME = "Giornaliero"
DAY_COLUMNS = list("BCDEFGHIJKLM")  # 12 giorni
FIRST_ROW = 4
LAST_ROW = 11        # attività/descrizione/logistica del giorno
DINNER_ROW = 12       # sempre la cena, per ogni colonna
EVENING_ROW = 13      # sempre la nota/attività serale, per ogni colonna
YEAR = 2026
MONTH = 8

HOME_DINNER_PATTERN = re.compile(r"home|a casa|casa\b", re.IGNORECASE)

ICONS = [
    (r"malpensa|\bjsi\b|\bmpx\b|volo", "✈️"),      # ✈️ voli/treni
    (r"bagagli", "\U0001F9F3"),                               # 🧳
    (r"\bspesa\b", "\U0001F6D2"),                              # 🛒
    (r"dinner|cena|ristorante|restaurant|basilikos|windmill", "\U0001F37D️"),  # 🍽️
    (r"escursione|barca|boat", "⛵"),                      # ⛵
    (r"motorino|scooter", "\U0001F6F5"),                       # 🛵
    (r"arrivo|consegna", "\U0001F6EC"),                        # 🛬
]
DEFAULT_ICON = "\U0001F4CD"  # 📍

WEEKDAYS_IT = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]

# owner della giornata (chi ha in mano il programma) — indicizzato per giorno del mese
OWNERS = {
    16: "Giulia", 17: "Giulia",
    18: "Giacomo", 19: "Giacomo",
    20: "Giulia", 21: "Giulia",
    22: "Giacomo", 23: "Giacomo", 24: "Giacomo",
}

# blocchi che erano solo un'annotazione informale di owner nel foglio originale
# (ora sostituita dal badge "owner" strutturato) -> li scartiamo per non duplicare
STRAY_NAME_ONLY = {"giulia", "giacomo"}

# luoghi/ristoranti riconosciuti nei testi -> query per Google Maps.
# Ordine: le frasi più specifiche/lunghe vanno PRIMA di quelle più generiche
# che le contengono (es. "Skopelos Town" prima di "Skopelos"), perché il
# regex di alternanza prova le alternative nell'ordine in cui compaiono.
PLACES = [
    ("Agios Ioannis a Castri", "Agios Ioannis Kastri, Skopelos, Grecia"),
    ("Monastero di Evangelistria", "Moni Evangelistrias, Skiathos, Grecia"),
    ("Holy Monastery of the Annunciation to the Virgin Mary", "Moni Evangelistrias, Skiathos, Grecia"),
    ("Evangelistria", "Moni Evangelistrias, Skiathos, Grecia"),
    ("The Windmill Restaurant", "The Windmill Restaurant, Skiathos, Grecia"),
    ("Skopelos Town", "Skopelos Town, Grecia"),
    ("Agia Eleni", "Agia Eleni Beach, Skiathos, Grecia"),
    ("Aegia Eleni", "Agia Eleni Beach, Skiathos, Grecia"),
    ("Kryfi Ammos", "Kryfi Ammos Beach, Skiathos, Grecia"),
    ("Krifi Ammos", "Kryfi Ammos Beach, Skiathos, Grecia"),
    ("Elia Beach", "Elia Beach, Skiathos, Grecia"),
    ("Lalaria Beach", "Lalaria Beach, Skiathos, Grecia"),
    ("Tripia Petra", "Tripia Petra, Skiathos, Grecia"),
    ("Megas Gialos", "Megas Gialos, Skiathos, Grecia"),
    ("Megas Giolas", "Megas Gialos, Skiathos, Grecia"),
    ("Mikros Aselinos", "Mikros Aselinos Beach, Skiathos, Grecia"),
    ("Raina Studio", "Raina Studio, Skiathos, Grecia"),
    ("Kastani", "Kastani Beach, Skopelos, Grecia"),
    ("Panormos", "Panormos Beach, Skopelos, Grecia"),
    ("Skotini", "Skotini Cave, Skiathos, Grecia"),
    ("Galazia", "Galazia Cave, Skiathos, Grecia"),
    ("Kastro", "Kastro, Skiathos, Grecia"),
    ("Skopelos", "Skopelos, Grecia"),
    ("Basilikos", "Basilikos, Skiathos, Grecia"),
    ("Elias", "Elia Beach, Skiathos, Grecia"),
]

# query Maps -> slug immagine condivisa (più frasi/variant possono puntare
# alla stessa query, es. "Skopelos" e "Skopelos Town" -> stessa foto)
QUERY_TO_SLUG = {
    "Agios Ioannis Kastri, Skopelos, Grecia": "agios-ioannis-kastri",
    "Moni Evangelistrias, Skiathos, Grecia": "moni-evangelistrias",
    "The Windmill Restaurant, Skiathos, Grecia": "windmill-restaurant",
    "Skopelos Town, Grecia": "skopelos",
    "Agia Eleni Beach, Skiathos, Grecia": "agia-eleni",
    "Kryfi Ammos Beach, Skiathos, Grecia": "kryfi-ammos",
    "Elia Beach, Skiathos, Grecia": "elia-beach",
    "Lalaria Beach, Skiathos, Grecia": "lalaria-beach",
    "Tripia Petra, Skiathos, Grecia": "tripia-petra",
    "Megas Gialos, Skiathos, Grecia": "megas-gialos",
    "Mikros Aselinos Beach, Skiathos, Grecia": "mikros-aselinos",
    "Raina Studio, Skiathos, Grecia": "raina-studio",
    "Kastani Beach, Skopelos, Grecia": "kastani",
    "Panormos Beach, Skopelos, Grecia": "panormos",
    "Skotini Cave, Skiathos, Grecia": "skotini-cave",
    "Galazia Cave, Skiathos, Grecia": "galazia-cave",
    "Kastro, Skiathos, Grecia": "kastro",
    "Skopelos, Grecia": "skopelos",
    "Basilikos, Skiathos, Grecia": "basilikos",
}

# etichetta leggibile mostrata sotto la foto nel carosello
SLUG_LABELS = {
    "agios-ioannis-kastri": "Agios Ioannis a Castri",
    "moni-evangelistrias": "Monastero di Evangelistria",
    "windmill-restaurant": "The Windmill Restaurant",
    "skopelos": "Skopelos",
    "agia-eleni": "Agia Eleni",
    "kryfi-ammos": "Kryfi Ammos",
    "elia-beach": "Elia Beach",
    "lalaria-beach": "Lalaria Beach",
    "tripia-petra": "Tripia Petra",
    "megas-gialos": "Megas Gialos",
    "mikros-aselinos": "Mikros Aselinos",
    "raina-studio": "Raina Studio",
    "kastani": "Kastani",
    "panormos": "Panormos",
    "skotini-cave": "Skotini Cave",
    "galazia-cave": "Galazia Cave",
    "kastro": "Kastro",
    "basilikos": "Basilikos",
}

IMAGES_DIR = WEBAPP_DIR / "images" / "places"

_PLACE_LOOKUP = {phrase.lower(): query for phrase, query in PLACES}
_PLACE_PATTERN = re.compile(
    r"\b(" + "|".join(re.escape(html.escape(p)) for p, _ in PLACES) + r")\b",
    re.IGNORECASE,
)
_PLACE_PATTERN_RAW = re.compile(
    r"\b(" + "|".join(re.escape(p) for p, _ in PLACES) + r")\b",
    re.IGNORECASE,
)


def find_image_slugs(raw_text):
    """Trova i luoghi con foto disponibile citati nel testo, in ordine di
    comparizione, senza duplicati."""
    if not raw_text:
        return []
    slugs = []
    seen = set()
    for m in _PLACE_PATTERN_RAW.finditer(raw_text):
        query = _PLACE_LOOKUP.get(m.group(0).lower())
        slug = QUERY_TO_SLUG.get(query)
        if slug and slug not in seen and (IMAGES_DIR / f"{slug}.jpg").exists():
            seen.add(slug)
            slugs.append(slug)
    return slugs


def linkify(escaped_text: str) -> str:
    def repl(m):
        matched = m.group(0)
        query = _PLACE_LOOKUP.get(html.unescape(matched).lower())
        if not query:
            return matched
        url = f"https://www.google.com/maps/search/?api=1&query={quote(query)}"
        return f'<a class="place-link" href="{url}" target="_blank" rel="noopener">{matched}</a>'

    return _PLACE_PATTERN.sub(repl, escaped_text)


def to_html(raw_text: str) -> str:
    escaped = html.escape(raw_text).replace("\n", "<br>")
    return linkify(escaped)


def guess_icon(text: str) -> str:
    low = text.lower()
    for pattern, icon in ICONS:
        if re.search(pattern, low):
            return icon
    return DEFAULT_ICON


def read_cell_text(ws, col: str, row: int):
    val = ws[f"{col}{row}"].value
    if val is None:
        return None
    text = unicodedata.normalize("NFC", str(val)).strip()
    if not text or text.lower() in STRAY_NAME_ONLY:
        return None
    return text


def classify_dinner(text):
    if not text:
        return {"status": "tbd", "html": None}
    status = "home" if HOME_DINNER_PATTERN.search(text) else "restaurant"
    return {"status": status, "html": to_html(text)}


def classify(text: str, bold: bool, size: float):
    stripped = text.strip()
    if len(stripped) >= 100:
        return "paragraph"
    if bold or (size and size >= 11):
        if len(stripped) <= 60:
            return "heading"
    return "note"


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Non trovo {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    days = []
    for i, col in enumerate(DAY_COLUMNS):
        day_num = 15 + i
        d = date(YEAR, MONTH, day_num)

        blocks = []
        for row in range(FIRST_ROW, LAST_ROW + 1):
            text = read_cell_text(ws, col, row)
            if not text:
                continue
            font = ws[f"{col}{row}"].font
            blocks.append({
                "text": text,
                "type": classify(text, bool(font.bold), font.size or 0),
            })

        # dedup: lo stesso testo può comparire più volte nel foglio
        # (es. celle segnaposto ripetute non ancora personalizzate)
        seen = set()
        deduped = []
        for b in blocks:
            key = b["text"].lower()
            if key in seen:
                continue
            seen.add(key)
            deduped.append(b)
        blocks = deduped

        daytime_texts = {b["text"].lower() for b in blocks}

        # riga 12 = cena, riga 13 = eventuale nota/attività serale: sempre
        # nella stessa posizione per ogni giornata, quindi le trattiamo
        # come sezione "Sera" separata invece che blocchi generici
        dinner_text = read_cell_text(ws, col, DINNER_ROW)
        evening_text = read_cell_text(ws, col, EVENING_ROW)
        if evening_text and evening_text.lower() in daytime_texts:
            # stesso testo già mostrato di giorno (es. placeholder ripetuto
            # come "sistemazione bagagli"): non duplicarlo nella sezione sera
            evening_text = None

        # foto dei luoghi citati in giornata (di giorno + cena + sera),
        # nell'ordine in cui compaiono, solo per i luoghi con un file
        # immagine disponibile in images/places/
        image_slugs = []
        for b in blocks:
            for slug in find_image_slugs(b["text"]):
                if slug not in image_slugs:
                    image_slugs.append(slug)
        for slug in find_image_slugs(dinner_text) + find_image_slugs(evening_text):
            if slug not in image_slugs:
                image_slugs.append(slug)
        images = [
            {"slug": s, "src": f"images/places/{s}.jpg", "caption": SLUG_LABELS.get(s, s)}
            for s in image_slugs
        ]

        for b in blocks:
            if b["type"] == "note":
                b["icon"] = guess_icon(b["text"])
            b["html"] = to_html(b["text"])
            del b["text"]

        dinner = classify_dinner(dinner_text)
        is_complete = bool(blocks) or dinner["status"] != "tbd" or bool(evening_text)

        days.append({
            "date": d.isoformat(),
            "weekday": WEEKDAYS_IT[d.weekday()],
            "day": day_num,
            "owner": OWNERS.get(day_num),
            "blocks": blocks,
            "isComplete": is_complete,
            "images": images,
            "evening": {
                "dinner": dinner,
                "activity": to_html(evening_text) if evening_text else None,
            },
        })

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps({"trip": "Skiathos", "year": YEAR, "days": days},
                          ensure_ascii=False, indent=2)
    OUT_PATH.write_text(payload, encoding="utf-8")
    print(f"Scritto {OUT_PATH} ({len(days)} giorni)")

    bump_sw_cache_version()


# tutti i file che il service worker mette in cache (tranne sw.js stesso):
# se cambia anche solo uno di questi, la versione della cache deve cambiare,
# altrimenti chi ha gia' installato la PWA continua a vedere codice vecchio
CACHED_ASSETS = [
    WEBAPP_DIR / "index.html",
    WEBAPP_DIR / "css" / "style.css",
    WEBAPP_DIR / "js" / "app.js",
    WEBAPP_DIR / "data" / "itinerary.json",
    WEBAPP_DIR / "manifest.webmanifest",
]


def bump_sw_cache_version():
    """Aggiorna CACHE_NAME in sw.js con un hash del contenuto di tutti gli
    asset in cache, cosi il service worker si accorge automaticamente che
    c'e' qualcosa di nuovo da scaricare (altrimenti chi ha gia' installato
    la PWA continuerebbe a vedere la versione vecchia)."""
    if not SW_PATH.exists():
        return
    hasher = hashlib.sha1()
    for path in CACHED_ASSETS:
        if path.exists():
            hasher.update(path.read_bytes())
    digest = hasher.hexdigest()[:8]
    sw_text = SW_PATH.read_text(encoding="utf-8")
    new_sw_text = re.sub(
        r'const CACHE_NAME = "skiathos-2026-[^"]*";',
        f'const CACHE_NAME = "skiathos-2026-{digest}";',
        sw_text,
        count=1,
    )
    if new_sw_text != sw_text:
        SW_PATH.write_text(new_sw_text, encoding="utf-8")
        print(f"Aggiornato {SW_PATH} (cache version {digest})")


if __name__ == "__main__":
    main()
